#!/usr/bin/env python3
"""
Cruzamento: CSV de processos Capano x Lista DEPRE completa.
Estratégia: Parse completo do PDF (~0.75s) + indexação + lookup O(1).

Matches por dois campos:
  - numero_processo_com_mascara (CSV) == Nº Processo DEPRE (PDF)
  - numero_processo_com_mascara (CSV) == Nº de Autos (PDF)

Duplicidades: quando um mesmo número do CSV gera múltiplas linhas DEPRE,
cada linha é incluída com flag de possível duplicidade.
"""

import subprocess
import re
import csv
import sys
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_full_text(pdf_path):
    """Extrai texto completo do PDF usando pdftotext."""
    print(f"Extraindo texto do PDF ({pdf_path})...")
    t0 = time.time()
    result = subprocess.run(
        ["pdftotext", "-layout", pdf_path, "-"],
        capture_output=True, text=True, timeout=600
    )
    if result.returncode != 0:
        raise RuntimeError(f"pdftotext falhou: {result.stderr}")
    t1 = time.time()
    print(f"  Texto extraído em {t1-t0:.2f}s ({len(result.stdout)/1024/1024:.1f}MB)")
    return result.stdout


def parse_all_orders(text):
    """Parse todas as ordens de pagamento do texto extraído."""
    print("Parseando ordens de pagamento...")
    t0 = time.time()

    blocks = re.split(r'(?=Ordem de Pagamento:\s*\d+)', text)
    orders = []

    for block in blocks:
        block = block.strip()
        if not block.startswith("Ordem de Pagamento:"):
            continue

        order = {}

        m = re.search(r'Ordem de Pagamento:\s*(\d+)', block)
        if m:
            order['ordem_pagamento'] = int(m.group(1))

        m = re.search(r'Nº Processo DEPRE:\s*([\d\.\-]+)', block)
        if m:
            order['processo_depre'] = m.group(1).strip()

        m = re.search(r'Natureza:\s+([\w\s]+?)(?:\s{2,}|ES/EP|$)', block)
        if m:
            order['natureza'] = m.group(1).strip()

        m = re.search(r'ES/EP:\s*([\d/]+)', block)
        order['es_ep'] = m.group(1).strip() if m else ''

        m = re.search(r'Nº de autos:\s+([\d\.\-]+)', block)
        if m:
            order['num_autos'] = m.group(1).strip()

        m = re.search(r'Ordem Orçamentária:\s*([\d/\-A-Za-z]+)', block)
        if m:
            order['ordem_orcamentaria'] = m.group(1).strip()

        m = re.search(r'Suspenso\?\s*([SN])', block)
        if m:
            order['suspenso'] = m.group(1).strip()

        m = re.search(r'Data do Protocolo:\s+([\d/]+(?:\s[\d:\.]+)?)', block)
        if m:
            date_str = m.group(1).strip()
            date_only = re.match(r'(\d{2}/\d{2}/\d{4})', date_str)
            order['data_protocolo'] = date_only.group(1) if date_only else date_str

        m = re.search(r'Nº do Protocolo Geral:\s*(\d+)', block)
        order['protocolo_geral'] = m.group(1).strip() if m else ''

        autos_antigos = []
        m = re.search(r'Nº de autos antigos:\s*(.+?)(?=Advogado|Devedora)', block, re.DOTALL)
        if m:
            for line in m.group(1).strip().split('\n'):
                val = line.strip()
                if val and not val.startswith('Advogado') and not val.startswith('Devedora'):
                    autos_antigos.append(val)
        order['autos_antigos'] = ' | '.join(autos_antigos) if autos_antigos else ''

        advogados = []
        m = re.search(r'Advogado\(s\):\s*(.+?)(?=Devedora:)', block, re.DOTALL)
        if m:
            for line in m.group(1).strip().split('\n'):
                val = line.strip()
                if val and not val.startswith('Devedora'):
                    advogados.append(val)
        order['advogados'] = ' | '.join(advogados) if advogados else ''

        m = re.search(r'Devedora:\s+(.+?)(?:\n|$)', block)
        if m:
            order['devedora'] = m.group(1).strip()

        if 'ordem_pagamento' in order:
            orders.append(order)

    t1 = time.time()
    print(f"  {len(orders)} ordens parseadas em {t1-t0:.2f}s")
    return orders


def build_indexes(orders):
    """Constrói índices para lookup O(1)."""
    depre_index = {}   # processo_depre -> [order_index]
    autos_index = {}   # num_autos -> [order_index]

    for idx, order in enumerate(orders):
        depre = order.get('processo_depre', '')
        autos = order.get('num_autos', '')
        if depre:
            depre_index.setdefault(depre, []).append(idx)
        if autos:
            autos_index.setdefault(autos, []).append(idx)

    return depre_index, autos_index


def load_csv(csv_path):
    """Carrega CSV e retorna lista de dicts."""
    rows = []
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def load_saldo(saldo_csv_path):
    """Carrega CSV de saldo DEPRE e retorna dict numero_depre -> {natureza, saldo}."""
    saldo = {}
    with open(saldo_csv_path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            num = row['numero_depre'].strip()
            try:
                val = float(row['saldo']) if row['saldo'] else 0.0
            except ValueError:
                val = 0.0
            saldo[num] = {
                'natureza_saldo': row.get('natureza', '').strip(),
                'saldo': val,
            }
    return saldo


def match_and_build(csv_rows, orders, depre_index, autos_index, saldo_dict=None):
    """
    Cruza CSV com DEPRE.
    Retorna lista de dicts com dados combinados.
    Duplica linhas quando há múltiplos matches.
    """
    saldo_dict = saldo_dict or {}
    results = []
    seen_pairs = set()  # (csv_num, ordem_pagamento) para detectar overlap

    for csv_row in csv_rows:
        num = csv_row['numero_processo_com_mascara']
        matched_indices = set()
        match_types = {}  # idx -> tipo de match

        # Match por Nº Processo DEPRE
        for idx in depre_index.get(num, []):
            matched_indices.add(idx)
            match_types[idx] = 'DEPRE'

        # Match por Nº de Autos
        for idx in autos_index.get(num, []):
            if idx in matched_indices:
                match_types[idx] = 'DEPRE+AUTOS'  # match duplo
            else:
                matched_indices.add(idx)
                match_types[idx] = 'AUTOS'

        if not matched_indices:
            continue

        # Detectar duplicidade: mais de 1 ordem para o mesmo número CSV
        is_multiple = len(matched_indices) > 1

        for idx in sorted(matched_indices):
            order = orders[idx]
            pair_key = (num, order.get('ordem_pagamento', 0))

            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)

            row = {
                # Dados do CSV
                'csv_numero_processo': num,
                'csv_classe': csv_row.get('classe', ''),
                'csv_orgao_julgador': csv_row.get('orgao_julgador', ''),
                'csv_situacao': csv_row.get('situacao_processo', ''),
                'csv_data_ultima_mov': csv_row.get('data_ultima_movimentacao', ''),
                'csv_parte_ativa': csv_row.get('parte_ativa', ''),
                'csv_parte_passiva': csv_row.get('parte_passiva', ''),
                # Dados do DEPRE
                'depre_ordem_pagamento': order.get('ordem_pagamento', ''),
                'depre_processo': order.get('processo_depre', ''),
                'depre_natureza': order.get('natureza', ''),
                'depre_es_ep': order.get('es_ep', ''),
                'depre_num_autos': order.get('num_autos', ''),
                'depre_ordem_orcamentaria': order.get('ordem_orcamentaria', ''),
                'depre_suspenso': order.get('suspenso', ''),
                'depre_data_protocolo': order.get('data_protocolo', ''),
                'depre_protocolo_geral': order.get('protocolo_geral', ''),
                'depre_autos_antigos': order.get('autos_antigos', ''),
                'depre_advogados': order.get('advogados', ''),
                'depre_devedora': order.get('devedora', ''),
                # Saldo DEPRE (do banco de dados)
                'saldo_depre': saldo_dict.get(order.get('processo_depre', ''), {}).get('saldo', ''),
                # Metadados do match
                'tipo_match': match_types.get(idx, ''),
                'possivel_duplicidade': 'SIM' if is_multiple else 'NÃO',
            }
            results.append(row)

    return results


def write_excel(results, output_path):
    """Gera Excel formatado com os dados cruzados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Capano x DEPRE"

    headers = [
        ('Nº Processo (CSV)', 30),
        ('Classe (CSV)', 25),
        ('Órgão Julgador (CSV)', 40),
        ('Situação (CSV)', 20),
        ('Última Mov. (CSV)', 18),
        ('Parte Ativa (CSV)', 30),
        ('Parte Passiva (CSV)', 30),
        ('Ordem Pagamento (DEPRE)', 18),
        ('Nº Processo DEPRE', 30),
        ('Natureza (DEPRE)', 20),
        ('ES/EP (DEPRE)', 15),
        ('Nº Autos (DEPRE)', 30),
        ('Ordem Orçamentária (DEPRE)', 22),
        ('Suspenso? (DEPRE)', 12),
        ('Data Protocolo (DEPRE)', 16),
        ('Protocolo Geral (DEPRE)', 18),
        ('Autos Antigos (DEPRE)', 30),
        ('Advogados (DEPRE)', 60),
        ('Devedora (DEPRE)', 45),
        ('Saldo DEPRE (R$)', 20),
        ('Tipo Match', 15),
        ('Possível Duplicidade', 18),
    ]

    field_keys = [
        'csv_numero_processo', 'csv_classe', 'csv_orgao_julgador',
        'csv_situacao', 'csv_data_ultima_mov', 'csv_parte_ativa',
        'csv_parte_passiva',
        'depre_ordem_pagamento', 'depre_processo', 'depre_natureza',
        'depre_es_ep', 'depre_num_autos', 'depre_ordem_orcamentaria',
        'depre_suspenso', 'depre_data_protocolo', 'depre_protocolo_geral',
        'depre_autos_antigos', 'depre_advogados', 'depre_devedora',
        'saldo_depre',
        'tipo_match', 'possivel_duplicidade',
    ]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    csv_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    depre_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    saldo_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    dup_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header row
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Dados (ordenados por ordem de pagamento decrescente)
    sorted_results = sorted(results, key=lambda r: (
        r.get('depre_ordem_pagamento', 0) if isinstance(r.get('depre_ordem_pagamento'), int)
        else int(r.get('depre_ordem_pagamento', 0) or 0)
    ), reverse=True)

    for row_idx, row_data in enumerate(sorted_results, 2):
        is_dup = row_data.get('possivel_duplicidade') == 'SIM'

        for col_idx, key in enumerate(field_keys, 1):
            value = row_data.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = thin_border

            # Formatar saldo como moeda
            if key == 'saldo_depre' and value != '':
                cell.number_format = '#,##0.00'

            # Colorir fundo por seção
            if is_dup:
                cell.fill = dup_fill
            elif col_idx <= 7:
                cell.fill = csv_fill
            elif col_idx == 20:  # Saldo DEPRE
                cell.fill = saldo_fill
            elif col_idx <= 19:
                cell.fill = depre_fill

    # Freeze + filter
    ws.freeze_panes = "A2"
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(sorted_results) + 1}"

    # Aba de resumo
    ws_summary = wb.create_sheet("Resumo")
    summary_data = [
        ("Métrica", "Valor"),
        ("Total processos CSV", len(set(r['csv_numero_processo'] for r in results))),
        ("Total matches encontrados", len(results)),
        ("Matches por Nº Processo DEPRE", sum(1 for r in results if 'DEPRE' in r['tipo_match'])),
        ("Matches por Nº de Autos", sum(1 for r in results if 'AUTOS' in r['tipo_match'])),
        ("Linhas com possível duplicidade", sum(1 for r in results if r['possivel_duplicidade'] == 'SIM')),
        ("Linhas sem duplicidade", sum(1 for r in results if r['possivel_duplicidade'] == 'NÃO')),
        ("Processos suspensos", sum(1 for r in results if r['depre_suspenso'] == 'S')),
        ("Processos ativos", sum(1 for r in results if r['depre_suspenso'] == 'N')),
        ("Saldo total DEPRE (R$)", sum(r.get('saldo_depre', 0) for r in results if isinstance(r.get('saldo_depre'), (int, float)))),
        ("Saldo ativos (R$)", sum(r.get('saldo_depre', 0) for r in results if isinstance(r.get('saldo_depre'), (int, float)) and r.get('depre_suspenso') == 'N')),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
        cell_a.border = thin_border
        cell_b.border = thin_border

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return len(sorted_results)


def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else "ListaPrecatorioPendente_707495.pdf"
    csv_path = sys.argv[2] if len(sys.argv) > 2 else "capano_tjsp_processos.csv"
    output_path = sys.argv[3] if len(sys.argv) > 3 else "reports/capano_x_depre.xlsx"
    saldo_path = sys.argv[4] if len(sys.argv) > 4 else "reports/depre_saldo.csv"

    total_t0 = time.time()

    # 1. Extrair texto completo do PDF
    text = extract_full_text(pdf_path)

    # 2. Parse todas as ordens
    orders = parse_all_orders(text)

    # 3. Indexar
    print("Construindo índices...")
    depre_index, autos_index = build_indexes(orders)
    print(f"  Índice DEPRE: {len(depre_index)} entradas")
    print(f"  Índice Autos: {len(autos_index)} entradas")

    # 4. Carregar CSV processos
    print(f"Carregando CSV ({csv_path})...")
    csv_rows = load_csv(csv_path)
    print(f"  {len(csv_rows)} processos carregados")

    # 5. Carregar saldo DEPRE
    saldo_dict = {}
    if os.path.exists(saldo_path):
        print(f"Carregando saldo DEPRE ({saldo_path})...")
        saldo_dict = load_saldo(saldo_path)
        print(f"  {len(saldo_dict)} saldos carregados")
    else:
        print(f"  AVISO: Arquivo de saldo não encontrado ({saldo_path})")

    # 6. Cruzar
    print("Cruzando dados...")
    results = match_and_build(csv_rows, orders, depre_index, autos_index, saldo_dict)
    print(f"  {len(results)} linhas no resultado")

    # 6. Gerar Excel
    print(f"Gerando Excel ({output_path})...")
    count = write_excel(results, output_path)

    total_t1 = time.time()
    print(f"\nConcluído em {total_t1-total_t0:.2f}s!")
    print(f"  {count} linhas exportadas para {output_path}")

    # Resumo
    dup_count = sum(1 for r in results if r['possivel_duplicidade'] == 'SIM')
    no_dup = sum(1 for r in results if r['possivel_duplicidade'] == 'NÃO')
    depre_match = sum(1 for r in results if 'DEPRE' in r['tipo_match'])
    autos_match = sum(1 for r in results if 'AUTOS' in r['tipo_match'])
    susp = sum(1 for r in results if r['depre_suspenso'] == 'S')
    ativo = sum(1 for r in results if r['depre_suspenso'] == 'N')

    print(f"\n--- RESUMO ---")
    saldo_total = sum(r.get('saldo_depre', 0) for r in results if isinstance(r.get('saldo_depre'), (int, float)))
    saldo_ativos = sum(r.get('saldo_depre', 0) for r in results if isinstance(r.get('saldo_depre'), (int, float)) and r.get('depre_suspenso') == 'N')
    com_saldo = sum(1 for r in results if isinstance(r.get('saldo_depre'), (int, float)) and r.get('saldo_depre', 0) > 0)

    print(f"  Matches por Nº Processo DEPRE: {depre_match}")
    print(f"  Matches por Nº de Autos:       {autos_match}")
    print(f"  Possíveis duplicidades:         {dup_count}")
    print(f"  Sem duplicidade:                {no_dup}")
    print(f"  Suspensos:                      {susp}")
    print(f"  Ativos:                         {ativo}")
    print(f"  Com saldo DEPRE > 0:            {com_saldo}")
    print(f"  Saldo total DEPRE:              R$ {saldo_total:,.2f}")
    print(f"  Saldo ativos:                   R$ {saldo_ativos:,.2f}")


if __name__ == "__main__":
    main()
