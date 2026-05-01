#!/usr/bin/env python3
"""
Extrator estruturado de dados de precatórios da Lista DEPRE (PDF).
Usa pdftotext para extração de texto e regex para parsing dos campos.
Gera arquivo Excel (.xlsx) com uma linha por ordem de pagamento.
"""

import subprocess
import re
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_text_from_pdf(pdf_path, start_page=1, end_page=None):
    """Extrai texto do PDF usando pdftotext com layout preservado."""
    cmd = ["pdftotext", "-layout", "-f", str(start_page)]
    if end_page:
        cmd.extend(["-l", str(end_page)])
    cmd.extend([pdf_path, "-"])

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    if result.returncode != 0:
        raise RuntimeError(f"pdftotext falhou: {result.stderr}")
    return result.stdout


def parse_orders(text):
    """Faz parsing do texto extraído e retorna lista de dicionários com os dados."""

    # Divide o texto em blocos por "Ordem de Pagamento:"
    blocks = re.split(r'(?=Ordem de Pagamento:\s*\d+)', text)

    orders = []
    for block in blocks:
        block = block.strip()
        if not block.startswith("Ordem de Pagamento:"):
            continue

        order = {}

        # Ordem de Pagamento
        m = re.search(r'Ordem de Pagamento:\s*(\d+)', block)
        if m:
            order['ordem_pagamento'] = int(m.group(1))

        # Nº Processo DEPRE
        m = re.search(r'Nº Processo DEPRE:\s*([\d\.\-]+)', block)
        if m:
            order['processo_depre'] = m.group(1).strip()

        # Natureza
        m = re.search(r'Natureza:\s+([\w\s]+?)(?:\s{2,}|ES/EP|$)', block)
        if m:
            order['natureza'] = m.group(1).strip()

        # ES/EP
        m = re.search(r'ES/EP:\s*([\d/]+)', block)
        if m:
            order['es_ep'] = m.group(1).strip()
        else:
            order['es_ep'] = ''

        # Nº de autos (primeiro, o principal)
        m = re.search(r'Nº de autos:\s+([\d\.\-]+)', block)
        if m:
            order['num_autos'] = m.group(1).strip()

        # Ordem Orçamentária
        m = re.search(r'Ordem Orçamentária:\s*([\d/\-A-Za-z]+)', block)
        if m:
            order['ordem_orcamentaria'] = m.group(1).strip()

        # Suspenso?
        m = re.search(r'Suspenso\?\s*([SN])', block)
        if m:
            order['suspenso'] = m.group(1).strip()

        # Data do Protocolo
        m = re.search(r'Data do Protocolo:\s+([\d/]+(?:\s[\d:\.]+)?)', block)
        if m:
            # Limpa timestamps longos, mantém só a data
            date_str = m.group(1).strip()
            date_only = re.match(r'(\d{2}/\d{2}/\d{4})', date_str)
            if date_only:
                order['data_protocolo'] = date_only.group(1)
            else:
                order['data_protocolo'] = date_str

        # Nº do Protocolo Geral
        m = re.search(r'Nº do Protocolo Geral:\s*(\d+)', block)
        if m:
            order['protocolo_geral'] = m.group(1).strip()
        else:
            order['protocolo_geral'] = ''

        # Nº de autos antigos (pode ter múltiplas linhas)
        autos_antigos = []
        m = re.search(r'Nº de autos antigos:\s*(.+?)(?=Advogado|Devedora)', block, re.DOTALL)
        if m:
            lines = m.group(1).strip().split('\n')
            for line in lines:
                val = line.strip()
                if val and not val.startswith('Advogado') and not val.startswith('Devedora'):
                    autos_antigos.append(val)
        order['autos_antigos'] = ' | '.join(autos_antigos) if autos_antigos else ''

        # Advogado(s) (pode ter múltiplos)
        advogados = []
        m = re.search(r'Advogado\(s\):\s*(.+?)(?=Devedora:)', block, re.DOTALL)
        if m:
            lines = m.group(1).strip().split('\n')
            for line in lines:
                val = line.strip()
                if val and not val.startswith('Devedora'):
                    advogados.append(val)
        order['advogados'] = ' | '.join(advogados) if advogados else ''

        # Devedora
        m = re.search(r'Devedora:\s+(.+?)(?:\n|$)', block)
        if m:
            order['devedora'] = m.group(1).strip()

        if 'ordem_pagamento' in order:
            orders.append(order)

    return orders


def write_excel(orders, output_path):
    """Gera arquivo Excel formatado com os dados extraídos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Precatórios DEPRE"

    # Headers
    headers = [
        ('Ordem Pagamento', 15),
        ('Nº Processo DEPRE', 30),
        ('Natureza', 20),
        ('ES/EP', 15),
        ('Nº de Autos', 30),
        ('Ordem Orçamentária', 22),
        ('Suspenso?', 12),
        ('Data Protocolo', 15),
        ('Nº Protocolo Geral', 20),
        ('Nº Autos Antigos', 30),
        ('Advogado(s)', 60),
        ('Devedora', 50),
    ]

    # Estilo do header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Dados (ordem invertida: mais recentes primeiro)
    orders = sorted(orders, key=lambda o: o.get('ordem_pagamento', 0), reverse=True)

    field_keys = [
        'ordem_pagamento', 'processo_depre', 'natureza', 'es_ep',
        'num_autos', 'ordem_orcamentaria', 'suspenso', 'data_protocolo',
        'protocolo_geral', 'autos_antigos', 'advogados', 'devedora'
    ]

    data_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, order in enumerate(orders, 2):
        for col_idx, key in enumerate(field_keys, 1):
            value = order.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(orders) + 1}"

    wb.save(output_path)
    return len(orders)


def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else "ListaPrecatorioPendente_707495.pdf"

    # Páginas para extrair (default: primeiras 5 para teste)
    start_page = int(sys.argv[2]) if len(sys.argv) > 2 else 1
    end_page = int(sys.argv[3]) if len(sys.argv) > 3 else 5

    output_path = sys.argv[4] if len(sys.argv) > 4 else "reports/precatorios_depre_teste.xlsx"

    # Garante diretório de saída
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    print(f"Extraindo texto do PDF (páginas {start_page}-{end_page})...")
    text = extract_text_from_pdf(pdf_path, start_page, end_page)

    print("Parseando ordens de pagamento...")
    orders = parse_orders(text)

    print(f"Encontradas {len(orders)} ordens de pagamento.")

    if not orders:
        print("ERRO: Nenhuma ordem encontrada. Verifique o PDF.")
        sys.exit(1)

    print(f"Gerando Excel em {output_path}...")
    count = write_excel(orders, output_path)

    print(f"Concluído! {count} ordens exportadas para {output_path}")

    # Print resumo das primeiras ordens para validação
    print("\n--- VALIDAÇÃO (primeiras ordens) ---")
    for o in orders[:3]:
        print(f"\nOrdem {o.get('ordem_pagamento')}:")
        for k, v in o.items():
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
